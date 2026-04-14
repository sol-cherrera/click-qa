"""
src/exporters/excel_exporter.py — Solo imágenes, sin textos ni encabezados.
Si el archivo está abierto en Excel, se usa xlwings para guardar desde la app (evita Permission denied).
"""
import io
import os
import tempfile
from datetime import datetime
from typing import List, Dict, Any

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from PIL import Image


# Ancho máximo embebido en Excel (px). Valores mayores = mejor detalle al ampliar (antes 840).
IMG_W      = 1920
ROW_H_PT   = 14.4   # alto de fila en puntos
PX_PER_ROW = 19.5   # px por fila (empirico, 100% zoom)
GAP_ROWS   = 2      # filas de separacion entre imagenes
# 72 pt = 1 inch, 96 px = 1 inch (Windows) → 1 px = 0.75 pt
PX_TO_PT   = 72 / 96


def export_excel(
    steps: List[Dict[str, Any]], filepath: str, sheet_name: str | None = None
) -> None:
    wb = None
    try:
        wb = load_workbook(filepath)
    except (PermissionError, OSError) as e:
        if getattr(e, "errno", None) == 13 or isinstance(e, PermissionError):
            # Archivo abierto en Excel: no podemos ni leer con openpyxl → usar xlwings
            _export_excel_via_xlwings(steps, filepath, requested_title=sheet_name)
            return
        raise
    except Exception:
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = _resolve_sheet_name(sheet_name, wb.sheetnames)
    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions["A"].width = 115
    current_row = 1
    _fill_sheet_openpyxl(ws, steps, current_row)

    try:
        wb.save(filepath)
    except (PermissionError, OSError) as e:
        if getattr(e, "errno", None) != 13 and not isinstance(e, PermissionError):
            raise
        # Guardar falló: archivo se abrió después → guardar vía xlwings
        _export_excel_via_xlwings(steps, filepath, requested_title=sheet_name)


def _sanitize_sheet_title(name: str) -> str:
    """Excel prohíbe []:*?/\\ en nombres de hoja; máximo 31 caracteres."""
    for c in "[]:*?/\\":
        name = name.replace(c, "")
    name = (name or "").strip() or "Capturas"
    return name[:31]


def _unique_sheet_title(base: str, existing: List[str]) -> str:
    base = _sanitize_sheet_title(base)
    if base not in existing:
        return base
    stem = base[:25].rstrip() or "Capturas"
    n = 2
    while True:
        suffix = f" ({n})"
        candidate = (stem + suffix)[:31]
        if candidate not in existing:
            return candidate
        n += 1


def _resolve_sheet_name(requested: str | None, existing: List[str]) -> str:
    if requested and str(requested).strip():
        return _unique_sheet_title(str(requested).strip(), existing)
    return _next_sheet_name(existing)


def suggested_export_sheet_title() -> str:
    """Texto sugerido para el nombre de hoja (puede coincidir; se ajusta al exportar)."""
    return f"Click {datetime.now().strftime('%d-%m %H.%M')}"[:31]


def _next_sheet_name(existing: List[str]) -> str:
    base_name = f"Click {datetime.now().strftime('%d-%m %H.%M')}"[:31]
    if base_name not in existing:
        return base_name
    base = base_name[:27]
    counter = 2
    while f"{base} ({counter})" in existing:
        counter += 1
    return f"{base} ({counter})"


def _fill_sheet_openpyxl(ws, steps: List[Dict[str, Any]], start_row: int) -> None:
    """Rellena la hoja con imágenes (openpyxl). No guarda el libro."""
    current_row = start_row
    for step in steps:
        if not step.get("screenshot_bytes"):
            continue
        try:
            buf, w, h = _prepare_image(step["screenshot_bytes"], IMG_W)
            xl_img = XlImage(buf)
            xl_img.width = w
            xl_img.height = h
            ws.add_image(xl_img, f"A{current_row}")
            rows_needed = max(int(h / PX_PER_ROW) + 2, 10)
            for r in range(current_row, current_row + rows_needed):
                ws.row_dimensions[r].height = ROW_H_PT
            current_row += rows_needed + GAP_ROWS
        except Exception:
            current_row += 1


def _export_excel_via_xlwings(
    steps: List[Dict[str, Any]], filepath: str, requested_title: str | None = None
) -> None:
    """
    Exporta usando xlwings: se conecta al libro ya abierto en Excel, añade hoja e imágenes, guarda.
    Así Excel guarda su propio archivo y no hay Permission denied.
    """
    import xlwings as xw

    path_abs = os.path.abspath(filepath)
    wb = xw.Book(path_abs)
    existing = [s.name for s in wb.sheets]
    sheet_name = _resolve_sheet_name(requested_title, existing)
    sheet = wb.sheets.add(name=sheet_name, after=wb.sheets[-1])
    sheet.range("A:A").column_width = 115

    current_row = 1
    temp_paths = []

    try:
        for step in steps:
            if not step.get("screenshot_bytes"):
                continue
            try:
                buf, w, h = _prepare_image(step["screenshot_bytes"], IMG_W)
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
                    f.write(buf.read())
                    tmp_path = f.name
                temp_paths.append(tmp_path)

                # Anchurar imagen en la celda (posición en puntos si hace falta)
                anchor = sheet.range(f"A{current_row}")
                sheet.pictures.add(
                    tmp_path, anchor=anchor,
                    width=w * PX_TO_PT, height=h * PX_TO_PT
                )

                rows_needed = max(int(h / PX_PER_ROW) + 2, 10)
                for r in range(current_row, current_row + rows_needed):
                    sheet.range(f"{r}:{r}").row_height = ROW_H_PT
                current_row += rows_needed + GAP_ROWS
            except Exception:
                current_row += 1

        wb.save(path_abs)
    finally:
        for p in temp_paths:
            try:
                os.unlink(p)
            except Exception:
                pass


def _prepare_image(screenshot_bytes: bytes, max_width: int):
    pil = Image.open(io.BytesIO(screenshot_bytes))
    if pil.width > max_width:
        ratio = max_width / pil.width
        pil = pil.resize((max_width, int(pil.height * ratio)), Image.LANCZOS)
    w, h = pil.size
    buf = io.BytesIO()
    pil.save(buf, format="PNG", compress_level=3)
    buf.seek(0)
    return buf, w, h
